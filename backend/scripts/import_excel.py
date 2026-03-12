#!/usr/bin/env python3
"""
Script para importar inventario desde Excel a Supabase.
Ejecutar UNA SOLA VEZ para migrar el Excel actual.

Uso:
    python scripts/import_excel.py --file "../inventario.xlsx"
    python scripts/import_excel.py --file "../inventario.xlsx" --sheet "Hoja1"
    python scripts/import_excel.py --file "../inventario.xlsx" --dry-run
"""

import argparse
import sys
import os
from pathlib import Path

# Agregar el directorio padre al path para importar app modules
sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    import openpyxl
except ImportError:
    print("❌ Instalar dependencia: pip install openpyxl")
    sys.exit(1)

from dotenv import load_dotenv
load_dotenv()

from supabase import create_client

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

# Posibles nombres de columnas (case insensitive)
COLUMN_MAPPINGS = {
    "sku":        ["sku", "codigo", "código", "code", "ref", "referencia", "id"],
    "name":       ["nombre", "name", "producto", "product", "descripcion", "descripción", "item"],
    "variety":    ["variedad", "variety", "tipo", "type", "modelo", "especificacion", "especificación"],
    "quantity":   ["cantidad", "qty", "quantity", "stock", "existencia", "existencias", "inventario"],
    "unit":       ["unidad", "unit", "ud", "medida"],
    "min_stock":  ["stock_minimo", "stock_mínimo", "min_stock", "minimo", "mínimo", "min", "stock min"],
}


def find_column(header_row: list, field: str) -> int | None:
    """Encuentra el índice de columna para un campo dado."""
    aliases = COLUMN_MAPPINGS.get(field, [])
    for idx, cell in enumerate(header_row):
        if cell and str(cell).strip().lower() in aliases:
            return idx
    return None


def parse_number(value) -> float:
    """Convierte un valor a float, devuelve 0 si no es posible."""
    if value is None:
        return 0.0
    try:
        return float(str(value).replace(",", ".").strip())
    except (ValueError, AttributeError):
        return 0.0


def generate_sku(name: str, idx: int) -> str:
    """Genera un SKU si no existe."""
    prefix = name[:3].upper().replace(" ", "") if name else "SKU"
    return f"{prefix}-{str(idx).zfill(4)}"


def import_from_excel(file_path: str, sheet_name: str = None, dry_run: bool = False):
    """Importa el inventario desde un archivo Excel."""

    print(f"\n🔍 Leyendo archivo: {file_path}")

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except FileNotFoundError:
        print(f"❌ Archivo no encontrado: {file_path}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Error al leer el archivo: {e}")
        sys.exit(1)

    # Seleccionar hoja
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(f"❌ Hoja '{sheet_name}' no encontrada. Hojas disponibles: {wb.sheetnames}")
            sys.exit(1)
        ws = wb[sheet_name]
    else:
        ws = wb.active
        print(f"📄 Usando hoja: {ws.title}")

    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        print("❌ El archivo está vacío")
        sys.exit(1)

    # Encontrar fila de encabezados
    header_row = None
    header_idx = 0
    for i, row in enumerate(rows):
        row_lower = [str(c).strip().lower() if c else "" for c in row]
        # Verificar si esta fila parece un encabezado (tiene palabras clave conocidas)
        all_aliases = [alias for aliases in COLUMN_MAPPINGS.values() for alias in aliases]
        matches = sum(1 for cell in row_lower if cell in all_aliases)
        if matches >= 2:
            header_row = [str(c).strip().lower() if c else "" for c in row]
            header_idx = i
            break

    if header_row is None:
        print("⚠️  No se encontró fila de encabezados automáticamente.")
        print(f"   Primera fila: {rows[0]}")
        print("   Asegúrate de que el Excel tenga encabezados como: SKU, Nombre, Variedad, Cantidad, Unidad, Stock_Minimo")
        sys.exit(1)

    print(f"\n📋 Encabezados encontrados en fila {header_idx + 1}: {header_row}")

    # Mapear columnas
    col_sku      = find_column(header_row, "sku")
    col_name     = find_column(header_row, "name")
    col_variety  = find_column(header_row, "variety")
    col_qty      = find_column(header_row, "quantity")
    col_unit     = find_column(header_row, "unit")
    col_min      = find_column(header_row, "min_stock")

    print(f"\n🗺️  Mapeo de columnas:")
    print(f"   SKU:         columna {col_sku + 1 if col_sku is not None else 'NO ENCONTRADA (se generará automáticamente)'}")
    print(f"   Nombre:      columna {col_name + 1 if col_name is not None else 'NO ENCONTRADA ⚠️'}")
    print(f"   Variedad:    columna {col_variety + 1 if col_variety is not None else 'no encontrada'}")
    print(f"   Cantidad:    columna {col_qty + 1 if col_qty is not None else 'NO ENCONTRADA ⚠️'}")
    print(f"   Unidad:      columna {col_unit + 1 if col_unit is not None else 'no encontrada'}")
    print(f"   Stock mín:   columna {col_min + 1 if col_min is not None else 'no encontrada'}")

    if col_name is None:
        print("\n❌ Columna 'Nombre/Producto' es obligatoria")
        sys.exit(1)

    # Procesar filas de datos
    items = []
    skipped = 0

    for i, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        if not row or all(c is None for c in row):
            continue

        name = str(row[col_name]).strip() if col_name is not None and row[col_name] else ""

        if not name or name.lower() in ["none", "nan", ""]:
            skipped += 1
            continue

        sku = str(row[col_sku]).strip() if col_sku is not None and row[col_sku] else generate_sku(name, len(items) + 1)
        variety = str(row[col_variety]).strip() if col_variety is not None and row[col_variety] else ""
        quantity = parse_number(row[col_qty]) if col_qty is not None else 0.0
        unit = str(row[col_unit]).strip() if col_unit is not None and row[col_unit] else "unidad"
        min_stock = parse_number(row[col_min]) if col_min is not None else 5.0

        item = {
            "sku": sku,
            "name": name,
            "variety": variety,
            "quantity": quantity,
            "unit": unit,
            "min_stock": min_stock,
        }
        items.append(item)

    print(f"\n✅ Items encontrados: {len(items)}")
    if skipped:
        print(f"⚠️  Filas vacías/saltadas: {skipped}")

    if not items:
        print("❌ No se encontraron items para importar")
        sys.exit(1)

    # Preview
    print("\n📦 Preview (primeros 5 items):")
    for item in items[:5]:
        print(f"   [{item['sku']}] {item['name']} - {item['variety']} | qty: {item['quantity']} {item['unit']} | min: {item['min_stock']}")

    if dry_run:
        print("\n🔵 DRY RUN - No se guardó nada en Supabase")
        print(f"   Se importarían {len(items)} items")
        return

    # Importar a Supabase
    print(f"\n📤 Importando {len(items)} items a Supabase...")

    if not SUPABASE_URL or not SUPABASE_KEY:
        print("❌ Variables SUPABASE_URL y SUPABASE_KEY no encontradas en .env")
        sys.exit(1)

    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

    success = 0
    errors = 0

    # Insertar en lotes de 50
    batch_size = 50
    for i in range(0, len(items), batch_size):
        batch = items[i:i + batch_size]
        try:
            result = supabase.table("inventory").upsert(
                batch,
                on_conflict="sku"  # Si el SKU ya existe, actualiza
            ).execute()
            success += len(batch)
            print(f"   ✅ Lote {i // batch_size + 1}: {len(batch)} items importados")
        except Exception as e:
            errors += len(batch)
            print(f"   ❌ Error en lote {i // batch_size + 1}: {e}")

    print(f"\n🎉 Importación completada!")
    print(f"   ✅ Exitosos: {success}")
    if errors:
        print(f"   ❌ Errores: {errors}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Importar inventario desde Excel a Supabase")
    parser.add_argument("--file", required=True, help="Ruta al archivo Excel (.xlsx)")
    parser.add_argument("--sheet", default=None, help="Nombre de la hoja (por defecto: hoja activa)")
    parser.add_argument("--dry-run", action="store_true", help="Solo mostrar preview, no importar")

    args = parser.parse_args()
    import_from_excel(args.file, args.sheet, args.dry_run)
